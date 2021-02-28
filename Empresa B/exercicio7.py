import numpy 
from openpyxl import Workbook
from openpyxl import load_workbook
from statistics import stdev

def criandoExcel():
    "Essa função inicia o arquivo excel"

    excelFile = Workbook()
    sheet = excelFile.active
    sheet.title = 'cadastro de alunos'

    #Colunas
    sheet.cell(row=1, column=1, value="DRE")
    sheet.cell(row=1, column=2, value="Curso")
    sheet.cell(row=1, column=3, value="Nome")
    sheet.cell(row=1, column=4, value="Gênero")
    sheet.cell(row=1, column=5, value="Data de nascimento")
    sheet.cell(row=1, column=6, value="Altura")
    sheet.cell(row=1, column=7, value="Peso")
    sheet.cell(row=1, column=8, value="CRA")
    sheet.cell(row=1, column=9, value="Créditos Obtidos")
    sheet.cell(row=1, column=10, value="Renda")

    return [excelFile,sheet]

def verificaExcelNovo(nome):
    "Essa função verifica se o excel já existe"
    existe = 0
    try:
        with open(nome, 'r') as f:
            existe =1
    except IOError:
        existe = 0

    return existe

def recupera(nome , nomeSheet):
    "Essa função define se vai ser usado algum arquivo existente ou um novo"
    
    existe = verificaExcelNovo(nome)
    print(existe)
    if existe==0: #arquivo não existe
        [excelFile, sheet] = criandoExcel()
    else:
        excelFile=load_workbook(filename = nome)
        sheet = excelFile[nomeSheet]

    return [excelFile,sheet]

def primeiraLinhaVazia(sheet, coluna):
    "Identifica primeira linha em branco no excel"

    count = 1
    while (not(sheet.cell(row=count, column=coluna).value is None)):
        count+=1
    return count

def primeiraColunaVazia(sheet, linha):
    "Identifica primeira coluna em branco no excel"
  
    count = 1
    while (not(sheet.cell(row=linha, column=count).value is None)):
        count+=1
    return count

def converteTabelaEmLista(sheet):
    "Transforma excel em lista"
    
    linhaVazia = primeiraLinhaVazia(sheet, 1)
    lista = []

    for linha in range(2,linhaVazia):
        dado=[]
        for coluna in range(1,11):
            dado.append(sheet.cell(row=linha, column=coluna).value)
        lista.append(dado)
    return lista

def dreUsado (lista, dre):
    "Verifica se o DRE já tá em uso"

    for i in range(len(lista)):
        if lista[i][0]==dre:
            return True
            break
    return False    

def adicionarDados(sheet):
    "Adiciona dado à planilha sheet"

    rotulos = ["DRE", "Curso", "Nome", "Gênero", "Data de Nascimento [DD/MM/AA]", "Altura [x.x]", "Peso [x.x]", "CRA", "Créditos Obtidos", "Renda [x,x]"]
    dados = []

    print("Insira os dados do novo cadastro")

    while True:
        try:
            for i in range(len(rotulos)):
                dado=input("%s: " % rotulos[i])
                dados.append(dado)
            [dados[5], dados[6], dados[7], dados[8], dados[9]]= [float(dados[5]), float(dados[6]), float(dados[7]), float(dados[8]), float(dados[9].replace(',','.'))]

        except ValueError:
            print("\nEntradas inválidas")
        else:
            break


    ultimaLinha=primeiraLinhaVazia(sheet, 1)
    lista = converteTabelaEmLista(sheet)

    if(lista != None):
        if (dreUsado(lista, dados[0])):
            repetirDRE = input("O dre já está sendo usado, tem certeza que quer inserir esse dado mesmo assim? [sim/não] ")

            if repetirDRE=='sim':
                for i in range(len(dados)):
                    sheet.cell(row=ultimaLinha, column=i+1, value=dados[i])
                return 0       
            else:
                return 1   
    for i in range(len(dados)):
        sheet.cell(row=ultimaLinha, column=i+1, value=dados[i])
    return 0

def editarDados(sheet):
    "Apagar dado da planilha sheet"

    lista = converteTabelaEmLista(sheet)
    opcao = input("Escolha um dado para localizar o registro que será alterado: [dre/nome] ")
    rotulos = ["DRE", "Curso", "Nome", "Gênero", "Data de Nascimento [DD/MM/AA]", "Altura [x.x]", "Peso [x.x]", "CRA", "Créditos Obtidos", "Renda [x,x]"]

    busca = ''

    if opcao == 'dre':
        busca = input('Digite o número do DRE: ')
        linha=1
        for i in range(len(lista)):
            linha +=1
            if lista[i][0] == busca:

                dados = []

                while True:
                    try:
                        for i in range(len(rotulos)):
                            dado=input("%s: " % rotulos[i])
                            dados.append(dado)
                        [dados[5], dados[6], dados[7], dados[8], dados[9]]= [float(dados[5]), float(dados[6]), float(dados[7]), float(dados[8]), float(dados[9].replace(',','.'))]

                    except ValueError:
                        print("\nEntradas inválidas")
                    else:
                        break

                if(lista != None):
                    if (dreUsado(lista, dados[0])):
                        repetirDRE = input("O dre já está sendo usado, tem certeza que quer inserir esse dado mesmo assim? [sim/não] ")

                        if repetirDRE=='sim':
                            for i in range(len(dados)):
                                sheet.cell(row=linha, column=i+1, value=dados[i])
                            return 0       
                        else:
                            return 1   
                for i in range(len(dados)):
                    sheet.cell(row=linha, column=i+1, value=dados[i])

                return 0

    elif opcao == 'nome':

        busca = input('Digite o nome do registro: ')
        linha=1
        for i in range(len(lista)):
            linha +=1
            if lista[i][2] == busca:

                dados = []

                while True:
                    try:
                        for i in range(len(rotulos)):
                            dado=input("%s: " % rotulos[i])
                            dados.append(dado)
                        [dados[5], dados[6], dados[7], dados[8], dados[9]]= [float(dados[5]), float(dados[6]), float(dados[7]), float(dados[8]), float(dados[9].replace(',','.'))]

                    except ValueError:
                        print("\nEntradas inválidas")
                    else:
                        break

                if(lista != None):
                    if (dreUsado(lista, dados[0])):
                        repetirDRE = input("O dre já está sendo usado, tem certeza que quer inserir esse dado mesmo assim? [sim/não] ")

                        if repetirDRE=='sim':
                            for i in range(len(dados)):
                                sheet.cell(row=linha, column=i+1, value=dados[i])
                            return 0       
                        else:
                            return 1   
                for i in range(len(dados)):
                    sheet.cell(row=linha, column=i+1, value=dados[i])

                return 0
    return 1

def apagarDados(sheet):
    "Apagar dado da planilha sheet"

    lista = converteTabelaEmLista(sheet)
    dado = input("Escolha um dado para localizar o registro que será apagado: [dre/nome] ")

    busca = ''

    if dado == 'dre':
        busca = input('Digite o número do DRE: ')
        linha=1
        for i in range(len(lista)):
            linha +=1
            if lista[i][0] == busca:
                sheet.delete_rows(linha, 1)
                return 0
    elif dado == 'nome':
        busca = input ('Digite o nome: ')
        linha=1
        for i in range(len(lista)):
            linha +=1
            if lista[i][0] == busca:
                sheet.delete_rows(linha, 1)
                return 0
    return 1

def contagemDados(sheet):
    
    rotulos = ["curso", "gênero"]
    lista = converteTabelaEmLista(sheet)

    opcao = input("Você gostaria de realizar a contagem por gênero ou curso? [curso/gênero] ")

    contagem = 0

    if opcao == "curso":
        dado = input("Digite o nome do curso que você deseja realizar a contagem: ")
        for i in range(len(lista)):
            if lista[i][1] == dado:
                contagem+=1
    elif opcao == "gênero":
        dado = input("Digite o gênero que você deseja realizar a contagem: ")
        for i in range(len(lista)):
            if lista[i][3] == dado:
                contagem+=1
    
    return [contagem, opcao, dado]


def mediaCRA(sheet):
    
    lista = converteTabelaEmLista(sheet)

    opcao = input("Para qual curso você gostaria de calcular a média de CRA? [nome_curso/todos]")

    contagem = 0
    somaCRA = 0
    
    if opcao == "todos":
        for i in range(len(lista)):
            contagem+=1
            somaCRA+=lista[i][7]
    else:
        for i in range(len(lista)):
            if lista[i][1] == opcao:
                contagem+=1
                somaCRA+=lista[i][7]
    
    media = somaCRA/contagem

    return [media, opcao]

def desvioPadrao(sheet):
    
    lista = converteTabelaEmLista(sheet)
    opcao = input("Para qual curso você gostaria de calcular o desvio padrão? [nome_curso/todos]")

    listaCRAS=[]
    
    if opcao == "todos":
        for i in range(len(lista)):
            listaCRAS.append(lista[i][7])
    else:
        for i in range(len(lista)):
            if lista[i][1] == opcao:
                listaCRAS.append(lista[i][7])

    desvio = stdev(listaCRAS)

    return [media, opcao]

def menu():

    print(" Operações: \n")
    
    print('''    (1) Adicionar dados\n
    (2) Alterar dados\n
    (3) Deletar dados\n
    (4) Contagem\n
    (5) Media CRA\n
    (6) Desvio Padrão\n''')

def interacaoUsuario():

    [excelFile,sheet] = recupera('cadastro-alunos.xlsx', 'cadastro de alunos')

    print("Cadastro de Alunos\n")

    menuPrint = menu()

    while True:
        try:
            opcao = int(input("Qual o número da opção desejada? "))
            if not (1<= opcao and opcao<=7):
                raise ValueError(opcao)
        except ValueError as e:
            print("Valor inválido")
        else:
            break

    sucesso = 1

    if opcao == 1:        
        sucesso = adicionarDados(sheet)
    elif opcao == 2:
        sucesso = editarDados(sheet)
    elif opcao == 3:
        sucesso = apagarDados(sheet)
    elif opcao == 4:
        [contagem, campo, dado] = contagemDados(sheet)
        print("O total de cadastros com o campo \"%s\" igual a \"%s\" é %d" % (campo, dado, contagem))
    elif opcao == 5:
        [media, curso] = mediaCRA(sheet)
        print("A media de CRA para os aluno do curso \"%s\" igual a %.2f" % (curso,media))
   elif opcao == 5:
        [media, curso] = mediaCRA(sheet)
        print("A media de CRA para os aluno do curso \"%s\" igual a %.2f" % (curso,media))

    
     
    if sucesso == 0:
        salvar = input("Gostaria de salvar sua alteração? [sim/não] ")
        if salvar == 'sim': 
            excelFile.save('cadastro-alunos.xlsx')
    
    novo = input("Gostaria de realizar uma outra análise ou edição? [sim/não] ")
    if novo == 'sim': 
        return interacaoUsuario()

interacaoUsuario()